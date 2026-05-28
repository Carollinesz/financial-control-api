import io

import openpyxl
import pytest

BASE = "/api/v1/transactions"
ACCOUNTS_BASE = "/api/v1/bank-accounts"


def _first_bank_id(client):
    return client.get("/api/v1/banks").json()[0]["bank_id"]


def _make_account(client):
    resp = client.post(ACCOUNTS_BASE, json={
        "bank_id": _first_bank_id(client),
        "account_name": "Test Account",
        "account_type": "checking",
        "start_value": 500.0,
    })
    return resp.json()["account_id"]


def _debit(account_id, description="Supermercado", value=-50.0):
    return {
        "account_id": account_id,
        "transaction_date": "2024-01-15",
        "value": value,
        "description": description,
        "category": "Alimentação",
        "type": "debit",
        "tracking": True,
    }


def _credit(account_id):
    return {
        "account_id": account_id,
        "transaction_date": "2024-01-15",
        "value": -1200.0,
        "description": "Notebook",
        "category": "Tecnologia",
        "type": "credit",
        "details": {
            "installments": 3,
            "first_payment": "2024-02-01",
            "interest": 0.0,
        },
        "tracking": True,
    }


# ── List ──────────────────────────────────────────────────────────────────────

def test_list_empty(client):
    assert client.get(BASE).status_code == 200
    assert client.get(BASE).json() == []


def test_list_returns_created(client):
    acc = _make_account(client)
    client.post(BASE, json=_debit(acc, "Padaria"))
    client.post(BASE, json=_debit(acc, "Farmácia"))
    data = client.get(BASE).json()
    descs = [t["description"] for t in data]
    assert "Padaria" in descs
    assert "Farmácia" in descs


# ── Create ────────────────────────────────────────────────────────────────────

def test_create_debit(client):
    acc = _make_account(client)
    resp = client.post(BASE, json=_debit(acc))
    assert resp.status_code == 201
    data = resp.json()
    assert data["type"] == "debit"
    assert data["description"] == "Supermercado"
    assert "transaction_id" in data


def test_create_credit(client):
    acc = _make_account(client)
    resp = client.post(BASE, json=_credit(acc))
    assert resp.status_code == 201
    data = resp.json()
    assert data["type"] == "credit"
    assert data["details"]["installments"] == 3


def test_create_account_not_found(client):
    assert client.post(BASE, json=_debit(999999)).status_code == 404


def test_create_credit_without_details_returns_400(client):
    acc = _make_account(client)
    resp = client.post(BASE, json={
        "account_id": acc,
        "transaction_date": "2024-01-15",
        "value": -500.0,
        "description": "Compra",
        "type": "credit",
    })
    assert resp.status_code == 400


def test_create_credit_without_first_payment_returns_400(client):
    acc = _make_account(client)
    resp = client.post(BASE, json={
        "account_id": acc,
        "transaction_date": "2024-01-15",
        "value": -500.0,
        "description": "Compra",
        "type": "credit",
        "details": {"installments": 2},
    })
    assert resp.status_code == 400


def test_create_credit_invalid_installments_returns_400(client):
    acc = _make_account(client)
    resp = client.post(BASE, json={
        "account_id": acc,
        "transaction_date": "2024-01-15",
        "value": -500.0,
        "description": "Compra",
        "type": "credit",
        "details": {"installments": 0, "first_payment": "2024-02-01"},
    })
    assert resp.status_code == 400


# ── Get ───────────────────────────────────────────────────────────────────────

def test_get_by_id(client):
    acc = _make_account(client)
    created = client.post(BASE, json=_debit(acc)).json()
    resp = client.get(f"{BASE}/{created['transaction_id']}")
    assert resp.status_code == 200
    assert resp.json()["transaction_id"] == created["transaction_id"]


def test_get_not_found(client):
    assert client.get(f"{BASE}/999999").status_code == 404


# ── Update ────────────────────────────────────────────────────────────────────

def test_update_description(client):
    acc = _make_account(client)
    created = client.post(BASE, json=_debit(acc)).json()
    resp = client.patch(f"{BASE}/{created['transaction_id']}", json={"description": "Padaria"})
    assert resp.status_code == 200
    assert resp.json()["description"] == "Padaria"


def test_update_value(client):
    acc = _make_account(client)
    created = client.post(BASE, json=_debit(acc)).json()
    resp = client.patch(f"{BASE}/{created['transaction_id']}", json={"value": -15.5})
    assert resp.status_code == 200
    assert float(resp.json()["value"]) == -15.5


def test_update_immutable_field_returns_422(client):
    acc = _make_account(client)
    created = client.post(BASE, json=_debit(acc)).json()
    resp = client.patch(f"{BASE}/{created['transaction_id']}", json={"type": "credit"})
    assert resp.status_code == 422


def test_update_not_found(client):
    assert client.patch(f"{BASE}/999999", json={"description": "X"}).status_code == 404


# ── Delete ────────────────────────────────────────────────────────────────────

def test_delete_success(client):
    acc = _make_account(client)
    created = client.post(BASE, json=_debit(acc)).json()
    assert client.delete(f"{BASE}/{created['transaction_id']}").status_code == 204
    assert client.get(f"{BASE}/{created['transaction_id']}").status_code == 404


def test_delete_not_found(client):
    assert client.delete(f"{BASE}/999999").status_code == 404


# ── Upload ────────────────────────────────────────────────────────────────────

def test_upload_unsupported_file_type(client):
    resp = client.post(
        f"{BASE}/upload",
        files={"file": ("data.csv", b"col1,col2\n1,2", "text/csv")},
    )
    assert resp.status_code == 415


def test_upload_xlsx_success(client):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["transaction_date", "value", "description"])
    ws.append(["2024-01-10", -30.0, "Café"])
    ws.append(["2024-01-11", -80.0, "Restaurante"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        f"{BASE}/upload",
        files={"file": ("transactions.xlsx", buf.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["created"] == 2
    assert data["errors"] == []


def test_upload_xlsx_missing_required_column(client):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["transaction_date", "description"])  # missing 'value'
    ws.append(["2024-01-10", "Café"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        f"{BASE}/upload",
        files={"file": ("transactions.xlsx", buf.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 422


def test_upload_xlsx_row_with_invalid_date(client):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["transaction_date", "value", "description"])
    ws.append(["not-a-date", -30.0, "Café"])
    ws.append(["2024-01-11", -80.0, "Valid row"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        f"{BASE}/upload",
        files={"file": ("transactions.xlsx", buf.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["created"] == 1
    assert len(data["errors"]) == 1
